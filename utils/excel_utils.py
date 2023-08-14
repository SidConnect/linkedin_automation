import openpyxl
import pandas as pd
from openpyxl import load_workbook


def read_excel_data(file_path, sheet_name):
    try:
        df = pd.read_excel(file_path, sheet_name)
        return df
    except Exception as e:
        raise Exception(f"Error occurred while reading data from excel: {str(e)}")


def write_to_excel(df, file_path, sheet_name):
    try:
        # Load the existing Excel workbook
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]

        # Find the next empty row
        next_row = sheet.max_row + 1

        # Write column names as the first row
        if next_row == 2:
            col_names = df.columns
            for col_index, col_name in enumerate(col_names, start=1):
                sheet.cell(row=1, column=col_index, value=col_name)

        # Write data to the next empty row
        for index, row in df.iterrows():
            for col_index, value in enumerate(row, start=1):
                sheet.cell(row=next_row, column=col_index, value=value)
            next_row += 1

        # Save the workbook
        workbook.save(file_path)
    except FileNotFoundError:
        # If the file doesn't exist, create a new one
        df.to_excel(file_path, sheet_name=sheet_name, index=False)


